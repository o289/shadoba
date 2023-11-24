import json
import pandas as pd
from openpyxl import load_workbook
from flask import Flask, render_template, request, send_file
from bs4 import BeautifulSoup

app = Flask(__name__)


with open('shadoba.json', 'r') as f:
    data = json.load(f)


root_html = 'index.html'
response_html = 'response.html'

@app.route('/')
def root():
    return render_template(root_html)

@app.route('/response', methods={'post'})
def response():
    rf = request.form

    format = rf['format']
    get_value = rf['get_value']
    card_type = rf['card_type']    
    count = 0
    
    source = data['shadoba_cards']
    
    
    if card_type == 'normal':
        cards = source[0]['cards']    
        
        if get_value == 'all':
            get_value = len(cards)
        else:
            get_value = int(get_value)

        if get_value <= -1:
            raise ValueError('0以上の数字にしてください。')

        
        # カード
        if format == 'card':
            
            created_html = '<div class="row">'
            
            for card in cards:
                if count == get_value:
                    break

                character = card['card']

                url = character['URL']
                created_html += f'<div class="col-lg-6  col-12 mx-auto my-4"><div class="card">'

                # キャラ名
                name = character['name']
                created_html += f"<div class='card-header'><h2 class='card-title text-center'>{name}</h2></div>"


                created_html += "<div class='card-body'>"

                # パック名
                pac = character['pac']
                created_html += f"<div class='card-title text-center'><h4>パック:{pac}</h4></div>"

                # レビュー
                reviews = character['reviews']
                try:
                    review = '<table class="table">'
                    for rev in reviews:
                        tr = rev.split(':')
                        th = tr[0]
                        td = tr[1]
                        review += f"<tr><th>{th}</th><td>{td}</td></tr>"

                    review += '</table>'
                    
                    created_html += f"<p class='card-text'>{review}</p>"
                except:
                    pass
                
                # 説明
                ability = character['ability']
                if 'null' in ability:
                    text_list = ''
                else:
                    created_html += '<h4 class="text-center pt-3">スキル</h4>'
                    text_list = '<ul>'
                    
                    for text in ability:
                        text_list += f"<li>{text}</li>"
                    
                    text_list += '</ul>'
                
                created_html += f"<p class='card-text'>{text_list}</p>"
                    
                
                created_html += '</div>'
                
                created_html += f'<div class="card-footer text-muted"><a href="{url}" class="d-block  w-50 mx-auto card-link  btn btn-danger">このキャラの情報はこちら</a></div>'

                created_html += '</div></div>'
                
                count += 1

            created_html += '</div>'
            
            
            soup = BeautifulSoup(created_html, 'html.parser')
            created_html = soup.prettify()
            
            return render_template(response_html, result=created_html)
        

        # エクセル
        elif format == 'excel':
            shadoba_excel = 'shadoba.xlsx'
            count = 0

            character_list = []
            

            for card in cards:
                if count == get_value:
                    break
                else:
                    character_data = {}
                    character = card['card']

                    name = character['name']
                    
                    url = character['URL']

                    pac = character['pac']


                    character_data['キャラ名'] = name
                    character_data['URL'] = url
                    character_data['パック'] = pac

                    reviews = character['reviews']
                    for rev in reviews:
                        tr = rev.split(':')
                        th = tr[0]
                        td = tr[1]
                        character_data[th] = td
                    
                    
                    ability = character['ability']
                    description_count = 1
                    
                    if 'null' in ability:
                        character_data[f'特殊能力{description_count}'] = 'なし'
                    else:
                        for text in ability:
                            character_data[f'特殊能力{description_count}'] = text
                            description_count += 1 
                            if not text in ability:
                                character_data[f'特殊能力{description_count}'] = 'なし'
                    

                    
                    character_list.append(character_data)
                    count += 1

            df = pd.DataFrame(character_list)
            df.to_excel(shadoba_excel, index=False)

            wb = load_workbook(shadoba_excel)
            ws = wb.worksheets[0]

            ws.column_dimensions['A'].width = 40
            ws.column_dimensions['B'].width = 60
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 20
            ws.column_dimensions['H'].width = 20
            ws.column_dimensions['J'].width = 20
            ws.column_dimensions['K'].width = 20
            ws.column_dimensions['L'].width = 100
            ws.column_dimensions['M'].width = 20
            ws.column_dimensions['N'].width = 100
            ws.column_dimensions['O'].width = 100
            ws.column_dimensions['P'].width = 100
            ws.column_dimensions['Q'].width = 100

            wb.save(shadoba_excel)

                
            return  send_file(shadoba_excel, as_attachment=True, download_name=shadoba_excel)
    elif card_type == 'additional':
        additional_cards = source[1]['additional_cards']

        if get_value == 'all':
            get_value = len(additional_cards)
        else:
            get_value = int(get_value)

        if get_value <= -1:
            raise ValueError('0以上の数字にしてください。')

        if format == 'card':
            
            created_html = '<div class="row">'
            
            for card in additional_cards:
                if count == get_value:
                    break

                character = card['additional_card']

                url = character['URL']
                created_html += f'<div class="col-lg-6  col-12 mx-auto my-4"><div class="card">'

                # キャラ名
                name = character['name']
                created_html += f"<div class='card-header'><h2 class='card-title text-center'>{name}</h2></div>"


                created_html += "<div class='card-body'>"

                # レビュー
                reviews = character['reviews']
    
                review = '<table class="table">'
                
                for rev in reviews:
                    tr = rev.split(':')
                    th = tr[0]
                    td = tr[1]
                    review += f"<tr><th>{th}</th><td>{td}</td></tr>"

                review += '</table>'
                
                created_html += '<h4 class="text-center">評価</h4>'
                created_html += f"<p class='card-text'>{review}</p>"

                # 能力
                ability = character['ability']
                if 'null' in ability:
                    text_list = ''
                else:
                    created_html += '<h4 class="text-center pt-3">スキル</h4>'
                    text_list = '<ul>'
                    
                    for text in ability:
                        text_list += f"<li>{text}</li>"
                    
                    text_list += '</ul>'
                
                created_html += f"<p class='card-text'>{text_list}</p>"

                
                created_html += '</div>'
                
                created_html += f'<div class="card-footer text-muted"><a href="{url}" class="d-block  w-50 mx-auto card-link  btn btn-primary">このキャラの情報はこちら</a></div>'

                created_html += '</div></div>'
                
                count += 1

            created_html += '</div>'
            
            
            soup = BeautifulSoup(created_html, 'html.parser')
            created_html = soup.prettify()
            
            return render_template(response_html, result=created_html)
        
        elif format == 'excel':
            shadoba_excel = 'shadoba.xlsx'
            count = 0

            character_list = []

            for card in additional_cards:
                if count == get_value:
                    break
                else:
                    character_data = {}
                    character = card['additional_card']

                    name = character['name']
                    
                    url = character['URL']

                    ability = character['ability']

                    character_data['キャラ名'] = name
                    character_data['URL'] = url

                    reviews = character['reviews']
                    for rev in reviews:
                        tr = rev.split(':')
                        th = tr[0]
                        td = tr[1]
                        character_data[th] = td
                    
                    description_count = 1

                    if 'null' in ability:
                        character_data[f'特殊能力{description_count}'] = 'なし'
                    else:
                        for text in ability:
                            if not text in ability:
                                character_data[f'特殊能力{description_count}'] = 'なし'
                            else:
                                character_data[f'特殊能力{description_count}'] = text
                            description_count += 1 
                            
                    
                    character_list.append(character_data)
                    count += 1

            df = pd.DataFrame(character_list)
            df.to_excel(shadoba_excel, index=False)

            wb = load_workbook(shadoba_excel)
            ws = wb.worksheets[0]

            ws.column_dimensions['A'].width = 40
            ws.column_dimensions['B'].width = 60
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 20
            ws.column_dimensions['H'].width = 20
            ws.column_dimensions['I'].width = 140
            ws.column_dimensions['J'].width = 140
            ws.column_dimensions['K'].width = 140
            ws.column_dimensions['L'].width = 140
            ws.column_dimensions['M'].width = 140
            ws.column_dimensions['N'].width = 150
            ws.column_dimensions['O'].width = 140
            ws.column_dimensions['P'].width = 140
            ws.column_dimensions['Q'].width = 20
            ws.column_dimensions['R'].width = 20
            ws.column_dimensions['S'].width = 20
            ws.column_dimensions['T'].width = 20

            wb.save(shadoba_excel)

                
            return  send_file(shadoba_excel, as_attachment=True, download_name=shadoba_excel)

        
        
    
    else:
        SyntaxError('想定外の値が挿入されました。')

